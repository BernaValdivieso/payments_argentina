from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for, current_app
import os
import pandas as pd
from werkzeug.utils import secure_filename
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings

main = Blueprint('main', __name__)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_excel(file_path):
    try:
        # Try to read the Excel file with Spanish sheet name first
        try:
            df = pd.read_excel(file_path, sheet_name='9. Remuneracion', engine='openpyxl')
            # Spanish column names
            costo_piezas_col = 'Costo Piezas'
            costo_horas_col = 'Costo Horas ($)'
            costo_total_col = 'Costo Total ($)'
        except:
            # If Spanish fails, try English sheet name
            df = pd.read_excel(file_path, sheet_name='9. Payroll', engine='openpyxl')
            # English column names
            costo_piezas_col = 'Cost per pieces ($)'
            costo_horas_col = 'Cost per hours ($)'
            costo_total_col = 'Total cost ($)'
        
        # Create a new workbook
        wb = Workbook()
        sheet = wb.active
        sheet.title = '9. Remuneracion'
        
        # Copy headers and apply formatting
        for col_idx, header in enumerate(df.columns, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            # Apply header formatting
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Get column indices
        costo_piezas_col_idx = df.columns.get_loc(costo_piezas_col) + 1
        costo_horas_col_idx = df.columns.get_loc(costo_horas_col) + 1
        costo_total_col_idx = df.columns.get_loc(costo_total_col) + 1
        
        # Copy data and update Costo Total
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                # Apply number format for currency columns
                if 'Costo' in df.columns[col_idx-1] or 'Cost' in df.columns[col_idx-1]:
                    cell.number_format = '#,##0.00'
            
            # Update Costo Total with the highest value
            costo_piezas = row[df.columns.get_loc(costo_piezas_col)]
            costo_horas = row[df.columns.get_loc(costo_horas_col)]
            costo_total = max(costo_piezas, costo_horas)
            sheet.cell(row=row_idx, column=costo_total_col_idx, value=costo_total)
        
        # Add new column for payment type
        tipo_pago_col = len(df.columns) + 1
        sheet.cell(row=1, column=tipo_pago_col, value='Tipo de Pago')
        # Format the new header
        header_cell = sheet.cell(row=1, column=tipo_pago_col)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')
        header_cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Process payment type
        for row_idx, row in enumerate(df.values, 2):
            costo_piezas = row[df.columns.get_loc(costo_piezas_col)]
            costo_horas = row[df.columns.get_loc(costo_horas_col)]
            
            if costo_piezas > costo_horas:
                sheet.cell(row=row_idx, column=tipo_pago_col, value='Piezas')
            else:
                sheet.cell(row=row_idx, column=tipo_pago_col, value='Horas')
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add auto-filter to all columns
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        
        return wb
        
    except Exception as e:
        flash(f'Error processing file: {str(e)}')
        raise

@main.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            processed_filename = f'processed_{timestamp}_{filename}'
            
            # Save uploaded file
            upload_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
            file.save(upload_path)
            
            try:
                # Process the file
                wb = process_excel(upload_path)
                
                # Save processed file
                processed_path = os.path.join(current_app.config['PROCESSED_FOLDER'], processed_filename)
                wb.save(processed_path)
                
                # Clean up uploaded file
                os.remove(upload_path)
                
                return send_file(processed_path, as_attachment=True)
            except Exception as e:
                flash(f'Error processing file: {str(e)}')
                return redirect(request.url)
        
        flash('Invalid file type. Please upload an Excel file.')
        return redirect(request.url)
    
    return render_template('index.html') 